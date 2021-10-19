from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import os

def get_fixed_weights():
    cell_weights = {}
    image_weights = {}

    for x in range(19):
        cell_weights[x] = 0.7
        image_weights[x] = 0.3

    return cell_weights, image_weights

def get_custom_weights():

    cell_weights = {}
    image_weights = {}

    for x in [0, 11, 16, 18]:
        cell_weights[x] = 1.0
        image_weights[x] = 0.0

    for x in [1, 14]:
        cell_weights[x] = 0.75
        image_weights[x] = 0.25

    for x in [2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17]:
        cell_weights[x] = 0.65
        image_weights[x] = 0.35

    for x in [13]:
        cell_weights[x] = 0.6
        image_weights[x] = 0.4

    return cell_weights, image_weights

def merge_solution_values(solution_values_list, solution_values_weights):
    result_solution_values = SolutionValues()

    for i in range(len(solution_values_list)):
        solution_values = solution_values_list[i]
        weight = solution_values_weights[i]

        for ID, num, label in solution_values.values:
            value = solution_values.get_value(ID, num, label) * weight
            result_solution_values.add_value(ID, num, label, value)

    return result_solution_values

class SolutionValues():

    def __init__(self):
        self.values = {}

    def add_value(self, ID, num, label, value):
        key = (ID, num, label)
        self.values[key] = self.values.get(key, 0) + value

    def get_value(self, ID, num, label):
        key = (ID, num, label)
        return self.values[key]

    def weight_cells_per_image(self, cell_weights, image_weights, border_and_garbage_value):
        temp_values = {}

        for ID, num, label in self.values:
            if label in [11, 18] or border_and_garbage_value[ID + '_' + str(num)] != 1.0:
                continue

            key = (ID, label)
            value = self.values[(ID, num, label)]

            if not key in temp_values:
                temp_values[key] = []

            temp_values[key].append(value)

        for ID, num, label in self.values:
            key = (ID, label)
            if label in [11, 18] or not key in temp_values:
                continue

            value = self.values[(ID, num, label)] * cell_weights[label]
            value += sum(temp_values[key]) / len(temp_values[key]) * image_weights[label]

            self.values[(ID, num, label)] = value

    def calculate_negatives(self):
        temp_values = {}

        for ID, num, label in self.values:
            key = (ID, num)
            value = self.values[(ID, num, label)]

            if not key in temp_values:
                temp_values[key] = []

            temp_values[key].append(value)

        for ID, num in temp_values:
            key = (ID, num)
            values = temp_values[(ID, num)]

            self.values[(ID, num, 18)] = 1.0 - max(values)

    def weight_border_and_garbage_images(self, border_and_garbage_value):
        for ID, num, label in self.values:
            self.values[(ID, num, label)] = self.values[(ID, num, label)] * border_and_garbage_value[ID + '_' + str(num)]

    def get_values_per_image(self):
        values_per_image = {}

        for ID, num, label in self.values:
            key = (ID, num)
            value = self.values[(ID, num, label)]

            if not key in values_per_image:
                values_per_image[key] = []

            values_per_image[key].append((label, value))

        return values_per_image

    def to_output_table(self, output_filename, output_sheet, append, short):

        if short is None or short < 0:
            short = 0

        values = { (key[0] + '_' + str(key[1]), key[2]) : value for key, value in self.values.items()}

        image_ids = sorted(set([(k[0], k[1]) for k, v in self.values.items()]))
        image_ids = [x[0] + '_' + str(x[1]) for x in image_ids]

        if append and os.path.isfile(output_filename):
            workbook = load_workbook(filename = output_filename)
            worksheet = workbook.create_sheet(output_sheet)
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = output_sheet

        for row, image_id in enumerate(image_ids, 2):
            worksheet.cell(row=row, column=1).value = image_id

        pf = PatternFill(fill_type='solid', start_color='FFBB77', end_color='FFBB77')
        pf2 = PatternFill(fill_type='solid', start_color='FFFF77', end_color='FFFF77')

        for row, image_id in enumerate(image_ids, 2):
            for label in range(19):
                value = values.get((image_id, label), -1)

                if value >= 0.5:
                    worksheet.cell(row=row, column=label+2).fill = pf
                elif value >= 0.1:
                    worksheet.cell(row=row, column=label+2).fill = pf2

                worksheet.cell(row=row, column=label+2).value = value

                if short > 0:
                    worksheet.cell(row=row, column=label+2).number_format = '0.' + '0' * short

        for index, column_cells in enumerate(worksheet.columns, 0):

            if index > 0 and short > 0:
                length = short + 2
            else:
                length = max(len(str(cell.value)) for cell in column_cells)

            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        for column in range(2,21):
            worksheet.cell(row=1, column=column).value = column-2

        workbook.save(filename = output_filename)
